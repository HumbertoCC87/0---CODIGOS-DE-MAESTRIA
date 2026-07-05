import pandas as pd
import re
import unicodedata
import nltk
import matplotlib.pyplot as plt
import numpy as np
import os
import sys
from nltk.corpus import stopwords

# Configurar stdout para soportar caracteres UTF-8 en terminales Windows
sys.stdout.reconfigure(encoding='utf-8')

from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.pipeline import Pipeline
from sklearn.naive_bayes import MultinomialNB
from sklearn.svm import LinearSVC
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix, ConfusionMatrixDisplay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle

# Crear carpetas necesarias
os.makedirs('matriz_confusion', exist_ok=True)
os.makedirs('salidas', exist_ok=True)
os.makedirs('salidas_excel', exist_ok=True)  # Carpeta para archivos Excel con formato

# Descargar las stopwords de NLTK si no las tienes
nltk.download('stopwords')

# ==================== FUNCIONES DE LIMPIEZA MEJORADAS ====================
def quitar_acentos(texto: str) -> str:
    """Elimina marcas de acento para normalizar palabras en español."""
    normalizado = unicodedata.normalize("NFD", texto)
    return "".join(caracter for caracter in normalizado if unicodedata.category(caracter) != "Mn")

def limpiar_texto(texto: str) -> str:
    """Normaliza texto: minúsculas, sin acentos, sin signos extraños y con espacios uniformes."""
    texto = str(texto).lower()
    texto = quitar_acentos(texto)
    texto = re.sub(r"http\S+|www\S+", " ", texto)
    texto = re.sub(r"[^a-zñ0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

# ==================== FUNCIÓN PARA APLICAR FORMATO A EXCEL ====================
def aplicar_formato_excel(df, nombre_archivo, tipo_dato):
    """Aplica formato de colores según el tipo de dato"""
    archivo_salida = f'salidas_excel/{nombre_archivo}.xlsx'
    
    # Limpiar caracteres de control ilegales para Excel
    df_clean = df.replace(to_replace=r'[\x00-\x08\x0b\x0c\x0e-\x1f]', value='', regex=True)
    
    # Guardar el DataFrame primero
    df_clean.to_excel(archivo_salida, index=True, engine='openpyxl')
    
    # Cargar el libro para aplicar formato
    wb = load_workbook(archivo_salida)
    ws = wb.active
    
    # Definir estilos según el tipo de dato
    if tipo_dato == 'top_terminos':
        # Colores para términos (azul) y scores (verde)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        termino_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        score_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Aplicar formato a columnas
        for row in range(2, ws.max_row + 1):
            # Columna A (término) - azul claro
            ws.cell(row, 1).fill = termino_fill
            # Columna B (score) - verde claro
            ws.cell(row, 2).fill = score_fill
        
    elif tipo_dato == 'similitud':
        from openpyxl.utils import get_column_letter
        # Escala de colores para similitud (valores entre 0 y 1)
        color_scale_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='FFC000',
                                        end_type='max', end_color='C00000')
        # Aplicar a todo el rango de datos
        if ws.max_row > 1:
            col_letter = get_column_letter(ws.max_column)
            ws.conditional_formatting.add(f'B2:{col_letter}{ws.max_row}', color_scale_rule)
        
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
    elif tipo_dato == 'bow_tf_tfidf':
        # Gradiente de color según el valor (term frequency)
        # Los valores más altos en rojo, más bajos en blanco
        for row in range(2, min(ws.max_row, 101)):  # Aplicar a primeras 100 filas
            for col in range(2, min(ws.max_column, 51)):  # Aplicar a primeras 50 columnas
                celda = ws.cell(row, col)
                if isinstance(celda.value, (int, float)) and celda.value > 0:
                    # Escala de colores según valor
                    intensidad = min(255, int(celda.value * 255))
                    if intensidad > 200:
                        celda.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif intensidad > 100:
                        celda.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    else:
                        celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        
    elif tipo_dato == 'procesamiento':
        # Colores para comentarios
        header_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        original_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
        procesado_fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
        categoria_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        
        # Aplicar formato a cada columna
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 1).fill = original_fill      # Comentario original
            ws.cell(row, 2).fill = procesado_fill     # Comentario procesado
            ws.cell(row, 3).fill = categoria_fill     # Categoría
    
    else:
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    
    # Formato común para encabezados
    if 'header_fill' in locals():
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(1, col)
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for celda in col:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(str(celda.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Guardar el archivo formateado
    wb.save(archivo_salida)
    print(f"✅ Archivo Excel formateado guardado: {archivo_salida}")

# ==================== CARGAR Y PREPROCESAR DATOS ====================
df = pd.read_csv('film_reviews_result.csv', sep='|')

generos_oficiales = [
    'Accion', 'Aventura', 'Ciencia Ficcion', 'Comedia', 'Drama', 
    'Fantasia', 'Musical', 'Romance', 'Terror', 'Western'
]

def asignar_genero_oficial(texto_genero):
    if pd.isna(texto_genero):
        return None
    texto_genero_limpio = limpiar_texto(texto_genero)
    
    for gen in generos_oficiales:
        if gen == 'Ciencia Ficcion' and ('ciencia ficcion' in texto_genero_limpio or 'scifi' in texto_genero_limpio):
            return gen
        elif gen.lower() in texto_genero_limpio:
            return gen
    return None

df['genero_limpio'] = df['gender'].apply(asignar_genero_oficial)
df_filtrado = df.dropna(subset=['genero_limpio']).copy()

print(f"Registros originales: {len(df)} -> Registros tras filtrar: {len(df_filtrado)}")
print("\nDistribución de géneros:")
print(df_filtrado['genero_limpio'].value_counts())

# Limpieza de texto
stop_words_es = list(stopwords.words('spanish'))
palabras_contexto = ['pelicula', 'peliculas', 'serie', 'series', 'director', 'actor', 'actores', 'ver', 'vista', 'pantalla']
stop_words_es.extend(palabras_contexto)

def limpiar_review(texto):
    if not isinstance(texto, str):
        return ""
    texto = limpiar_texto(texto)
    palabras = texto.split()
    palabras_filtradas = [w for w in palabras if w not in stop_words_es]
    return " ".join(palabras_filtradas)

df_filtrado['review_procesada'] = df_filtrado['review_text'].apply(limpiar_review)

# ==================== GUARDAR PROCESAMIENTO ====================
procesamiento_df = df_filtrado[['review_text', 'review_procesada', 'genero_limpio']].copy()
procesamiento_df.columns = ['comentario_original', 'comentario_procesado', 'categoria']
procesamiento_df.to_csv('salidas/procesamiento.csv', index=False, encoding='utf-8')
print("\n✅ Archivo CSV guardado: 'salidas/procesamiento.csv'")

# Crear Excel con formato para procesamiento
aplicar_formato_excel(procesamiento_df, 'procesamiento', 'procesamiento')

# ==================== REPRESENTACIONES NUMÉRICAS ====================
print("\n--- Generando representaciones numéricas ---")

# 1. Bag of Words
bow_vectorizer = CountVectorizer(max_features=1000, ngram_range=(1, 2))
X_bow = bow_vectorizer.fit_transform(df_filtrado['review_procesada'])
bow_df = pd.DataFrame(X_bow.toarray(), columns=[f'term_{i}' for i in range(X_bow.shape[1])])
bow_df.to_csv('salidas/Bow.csv', index=True)
print("✅ Bag of Words guardado (CSV)")

# Guardar una versión reducida en Excel (primeras 50 filas y 30 columnas para no sobrecargar)
bow_excel = bow_df.iloc[:50, :30]
aplicar_formato_excel(bow_excel, 'Bow', 'bow_tf_tfidf')

# 2. Matriz TF
tf_vectorizer = TfidfVectorizer(max_features=1000, ngram_range=(1, 2), use_idf=False)
X_tf = tf_vectorizer.fit_transform(df_filtrado['review_procesada'])
tf_df = pd.DataFrame(X_tf.toarray(), columns=[f'term_{i}' for i in range(X_tf.shape[1])])
tf_df.to_csv('salidas/matriz_tf.csv', index=True)
print("✅ Matriz TF guardada (CSV)")

tf_excel = tf_df.iloc[:50, :30]
aplicar_formato_excel(tf_excel, 'matriz_tf', 'bow_tf_tfidf')

# 3. Matriz TF-IDF
tfidf_vectorizer = TfidfVectorizer(max_features=1000, ngram_range=(1, 2), use_idf=True)
X_tfidf = tfidf_vectorizer.fit_transform(df_filtrado['review_procesada'])
tfidf_df = pd.DataFrame(X_tfidf.toarray(), columns=[f'term_{i}' for i in range(X_tfidf.shape[1])])
tfidf_df.to_csv('salidas/matriz_tf-idf.csv', index=True)
print("✅ Matriz TF-IDF guardada (CSV)")

tfidf_excel = tfidf_df.iloc[:50, :30]
aplicar_formato_excel(tfidf_excel, 'matriz_tfidf', 'bow_tf_tfidf')

# 4. Top términos TF-IDF
feature_names = tfidf_vectorizer.get_feature_names_out()
idf_scores = tfidf_vectorizer.idf_
idf_df = pd.DataFrame({'termino': feature_names, 'idf_score': idf_scores})
idf_df = idf_df.sort_values('idf_score', ascending=False)
idf_df.head(100).to_csv('salidas/top_terminos_tf-idf.csv', index=False)
print("✅ Top términos TF-IDF guardado (CSV)")

# Guardar Excel con formato especial
aplicar_formato_excel(idf_df.head(100), 'top_terminos_tfidf', 'top_terminos')

# 5. Similitud entre documentos
from sklearn.metrics.pairwise import cosine_similarity
X_tfidf_sample = X_tfidf[:100]
similitud_matrix = cosine_similarity(X_tfidf_sample)
similitud_df = pd.DataFrame(similitud_matrix)
similitud_df.columns = [f'doc_{i}' for i in range(len(similitud_df.columns))]
similitud_df.index = [f'doc_{i}' for i in range(len(similitud_df.index))]
similitud_df.to_csv('salidas/similitud_documentos.csv', index=True)
print("✅ Matriz de similitud guardada (CSV)")

# Guardar Excel con formato de gradiente
aplicar_formato_excel(similitud_df, 'similitud_documentos', 'similitud')

# ==================== DIVISIÓN Y ENTRENAMIENTO ====================
X = df_filtrado['review_procesada']
y = df_filtrado['genero_limpio']
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)

generos_presentes = y_test.unique()
print(f"\nGéneros presentes en prueba: {sorted(generos_presentes)}")

# Modelos
modelos = {
    'Naive Bayes Multinomial': MultinomialNB(),
    'Maquina de Soporte Vectorial (LinearSVC)': LinearSVC(random_state=42, max_iter=2000),
    'Regresion Logistica': LogisticRegression(max_iter=1000, random_state=42)
}

resultados = {}
print("\n--- Iniciando entrenamiento y evaluacion ---")

for nombre, algoritmo in modelos.items():
    pipeline = Pipeline([
        ('tfidf', TfidfVectorizer(max_features=5000, ngram_range=(1, 2))),
        ('clf', algoritmo)
    ])
    
    pipeline.fit(X_train, y_train)
    y_pred = pipeline.predict(X_test)
    acc = accuracy_score(y_test, y_pred)
    resultados[nombre] = acc
    
    print(f"\n{'='*60}")
    print(f"Resultados para: {nombre}")
    print(f"{'='*60}")
    print(f"Accuracy Global: {acc:.4f}")
    print("\nClassification Report:")
    print(classification_report(y_test, y_pred, zero_division=0))
    
    cm = confusion_matrix(y_test, y_pred, labels=generos_presentes)
    print("\nMatriz de Confusión:")
    print(pd.DataFrame(cm, index=generos_presentes, columns=generos_presentes))
    print(f"{'='*60}")
    
    # Guardar matriz de confusión
    cm_df = pd.DataFrame(cm, index=generos_presentes, columns=generos_presentes)
    cm_df.to_csv(f'matriz_confusion/matriz_confusion_{nombre.replace(" ", "_")}.csv', index=True)
    
    plt.figure(figsize=(10, 8))
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=generos_presentes)
    disp.plot(xticks_rotation=45, ax=plt.gca(), cmap='Blues', colorbar=True)
    plt.title(f'Matriz de Confusión - {nombre}', fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.savefig(f'matriz_confusion/matriz_confusion_{nombre.replace(" ", "_")}.png', dpi=300, bbox_inches='tight')
    plt.close()
    print(f"✅ Matriz guardada en 'matriz_confusion/'")

# Comparativa final
print("\n--- COMPARATIVA FINAL ---")
for nombre, acc in resultados.items():
    print(f"-> {nombre}: {acc:.4f}")

mejor_modelo = max(resultados, key=resultados.get)
print(f"\n🏆 Mejor modelo: ¡{mejor_modelo}!")

with open('salidas/resultados_modelos.txt', 'w', encoding='utf-8') as f:
    f.write("COMPARATIVA FINAL\n")
    f.write("="*40 + "\n")
    for nombre, acc in resultados.items():
        f.write(f"{nombre}: {acc:.4f}\n")
    f.write(f"\nMejor modelo: {mejor_modelo}\n")

print("\n✅ Proceso completado!")
print(f"   - Archivos CSV: carpeta 'salidas/'")
print(f"   - Archivos Excel con COLOR: carpeta 'salidas_excel/'")
print(f"   - Matrices de confusión: carpeta 'matriz_confusion/'")